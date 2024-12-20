from datetime import datetime

import os
from tkinter import ttk
import tkinter as tk
import win32com.client
import openpyxl
import pandas as pd

file_path = '데이터.xlsx'
file_path2 = r'C:\Users\user\dist\거래명세서및거래명세표.xlsx'
# Read the .xlsx file
df = pd.read_excel(file_path)

yous = df['거래처'].dropna().tolist()
mocks = df['품목'].tolist()
prices = df['단가'].tolist()
def save_data():
    # 입력값 가져오기
    you = you_var.get()  # StringVar를 사용하여 you_entry의 값 가져오기
    mock = mock_var.get() 
    num = int(num_entry.get())
    try:
        pay = int(pay_result.get())
    except (ValueError, AttributeError):
        pay = 0

    # XLSX 파일 이름과 경로 설정
    if you in yous:
        file_name = f'{you}거래서.xlsx'
    else:
        file_name = '기타거래서.xlsx'
    
    if mock in mocks:
        price = prices[mocks.index(mock)]  # 가격 설정

    tot2 = 0   
    tot = 0 
    totprice = price * num  # 공급가액 계산
    df2 = pd.read_excel(file_name) if file_name in os.listdir() else pd.DataFrame()  # 파일이 있으면 읽고, 없으면 빈 DataFrame 생성

    if not df2.empty:  # 파일이 있을 경우 공급가액 합산
        tot = df2['잔금'].iloc[-1]
        tot2 = tot + totprice
    else:
        tot2 = totprice  # 파일이 없으면 tot2는 totprice와 동일
    retot = tot2 - pay
    # jutot = df2['잔금'].iloc[-1] + totprice
    # 저장할 데이터 (날짜 포함)
    data = [you, mock, num, price, totprice, tot2, pay, retot, datetime.now().strftime('%Y-%m-%d')]

    # 컬럼명 설정 
    columns = ['거래처', '품목', '수량', '단가', '공급가액', '합계', '입금', '잔금', '거래날짜']

    # XLSX 파일이 이미 존재하는지 확인
    try:
        # 기존 엑셀 파일을 열기
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
    except FileNotFoundError:
        # 파일이 없다면 새로 생성
        wb = openpyxl.Workbook()
        sheet = wb.active
        # 헤더 추가 (첫 번째로 컬럼명 추가)
        sheet.append(columns)

    # 데이터 추가 (컬럼에 맞는 순서로 데이터가 들어가도록)
    sheet.append(data)

    # XLSX 파일에 저장
    wb.save(file_name)

    # 파일 저장 완료 메시지 출력 (옵션)
    print(f"{file_name}에 데이터가 저장되었습니다.")
    
    wb1 = openpyxl.load_workbook(file_path2)
    sheet = wb1.active
    sheet['G31'] = datetime.now().strftime('%Y-%m-%d')
    sheet['G35'] = you
    sheet['G37'] = totprice
    sheet['C40'] = mock
    sheet['R40'] = num
    sheet['V40'] = price
    sheet['Z40'] = totprice
    sheet['G50'] = tot
    sheet['Q50'] = tot2
    sheet['AB50'] = pay
    sheet['AB52'] = retot
    sheet['G4'] = datetime.now().strftime('%Y-%m-%d')
    sheet['G8'] = you
    sheet['G10'] = totprice
    sheet['C13'] = mock
    sheet['R13'] = num
    sheet['V13'] = price
    sheet['Z13'] = totprice
    sheet['G23'] = tot
    sheet['Q23'] = tot2
    sheet['AB23'] = pay
    sheet['AB25'] = retot
    wb1.save(file_path2)
    print_button.grid(row=3, column=1, padx=5, pady=5)

def startprint():
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    workbook = excel.Workbooks.Open(file_path2)
    workbook.PrintOut()
    excel.Quit()

def fine_data():
    fineyou = you_var.get()
    mock = mock_var.get() 
    num = int(num_entry.get())
    if fineyou in yous:
        find_name = f'{fineyou}거래서.xlsx'
    else:
        find_name = '기타거래서.xlsx'
    if mock in mocks:
        price = prices[mocks.index(mock)]  # 가격 설정
    totprice = price * num 
    df2 = pd.read_excel(find_name)
    tot2 = df2['잔금'].iloc[-1] + totprice

    fine_result.config(text=f"{tot2}") 
    return tot2
    
# GUI 생성
def pay_data():
    pay = int(pay_result.get())
    total = fine_data()
    result = total - pay
    result_result.config(text=f"{result}")
    you = you_var.get()

    # # XLSX 파일 이름과 경로 설정
    # if you in yous:
    #     file_name = f'{you}거래서.xlsx'
    # else:
    #     file_name = '기타거래서.xlsx'
    # df2 = pd.read_excel(file_name)
    # columns = ['거래처', '품목', '수량', '단가', '공급가액', '잔금', '거래날짜']
    # data = [you, None, None, None, None, result, datetime.now().strftime('%Y-%m-%d')]
    # try:
    #     # 기존 엑셀 파일을 열기
    #     wb = openpyxl.load_workbook(file_name)
    #     sheet = wb.active
    # except FileNotFoundError:
    #     # 파일이 없다면 새로 생성
    #     wb = openpyxl.Workbook()
    #     sheet = wb.active
    #     sheet.append(columns)
    # # 데이터 추가 (컬럼에 맞는 순서로 데이터가 들어가도록)
    # sheet.append(data)

    # # XLSX 파일에 저장
    # wb.save(file_name)

    # # 파일 저장 완료 메시지 출력 (옵션)
    # print(f"{file_name}에 데이터가 저장되었습니다.")


root = tk.Tk()
# StringVar로 OptionMenu에 선택된 값 관리


root.title('칼제비')
youoptions = yous
mockoptions = mocks

you_var = tk.StringVar()
you_var.set('')  # 기본값 설정

mock_var = tk.StringVar()
mock_var.set('')  # 기본값 설정
# 라벨 및 입력 필드 생성

you_label = ttk.Label(root, text='거래처')

you_entry = tk.OptionMenu(root, you_var, *youoptions)  # OptionMenu 사용

mock_label = ttk.Label(root, text='품목')

mock_entry = tk.OptionMenu(root, mock_var, *mockoptions)  # OptionMenu 사용

num_label = ttk.Label(root, text='수량')

num_entry = ttk.Entry(root)

# 저장 버튼 생성
fine_button = ttk.Button(root, text='조회', command=fine_data)
save_button = ttk.Button(root, text='저장', command=save_data)
print_button = ttk.Button(root, text='인쇄', command=startprint)
fine_label = ttk.Label(root, text='총액:')
fine_result = ttk.Label(root, text='　　　　　　　　')
pay_button = ttk.Button(root, text='지불', command=pay_data)
pay_result = ttk.Entry(root)
result_label = ttk.Label(root, text='잔액:')
result_result = ttk.Label(root, text='　　　　　　　　')

# 라벨 및 입력 필드 위치 지정
you_label.grid(row=0, column=0, padx=5, pady=5)
you_entry.grid(row=0, column=1, padx=5, pady=5)
fine_button.grid(row=0, column=2, padx=5, pady=5)
fine_label.grid(row=0, column=3, padx=5, pady=5)
fine_result.grid(row=0, column=4, padx=5, pady=5)

mock_label.grid(row=1, column=0, padx=5, pady=5)
mock_entry.grid(row=1, column=1, padx=5, pady=5)
pay_button.grid(row=1, column=2, padx=5, pady=5)
pay_result.grid(row=1, column=4, padx=5, pady=5)

num_label.grid(row=2, column=0, padx=5, pady=5)
num_entry.grid(row=2, column=1, padx=5, pady=5)
result_label.grid(row=2, column=3, padx=5, pady=5)
result_result.grid(row=2, column=4, padx=5, pady=5)

# 저장 버튼 위치 지정
save_button.grid(row=3, column=4, padx=5, pady=5)




root.mainloop()
